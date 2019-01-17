import os
import random
import timeit
import statistics
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from colorama import init
from termcolor import colored


"""
Outputs to .xlsx file
"""


class DataPoint(object):
    """
    Contains the information about each test run to use in results
    """
    def __init__(self, value1, value2, gcd, time):
        """
        :param value1: First value from random array pair
        :type: integer
        :param value2: Second value from random array pair
        :type: integer
        :param gcd: Calculated greatest comman divisor
        :type: integer
        :param time: Time ellapsed during GCD calculation
        :type: float
        """
        self.value1 = value1
        self.value2 = value2
        self.gcd = gcd
        self.time = time


# Iterate through all values of 'b', starting with 'b' and working down to 1
# If a number is a divisor of b, check if its also a divisor of 'a'
# Note 'b' is always <= a
def brute_force(a, b):
    i = b
    for n in range(1, b):
        if a % i == 0 and b % i == 0:
            return i
        i -= 1
    return 1


def original_euclid(a, b):
    remainder = None
    while 0 != remainder:
        quotient = a // b
        remainder = a - quotient * b
        a = b
        b = remainder
    return a


def second_euclid(a, b):
    remainder = None
    while 0 != remainder:
        remainder = a - b
        if remainder >= b:
            remainder = remainder - b
            if remainder >= b:
                remainder = remainder - b
                if remainder >= b:
                    remainder = a - b * (a // b)
        a = b
        b = remainder
    return a


def run_test(function, length, number_array):
    datapoints = []  # Array of Datapoint objects
    for i in range(0, length):
        # Gather random pair
        a = next(number_array)
        b = next(number_array)

        # preprocess
        if a <= 0 and b <= 0:
            return
        if a < b:
            tmp = a
            a = b
            b = tmp

        # Time GCD calc
        start_time = timeit.default_timer()
        gcd = function(a, b)
        stop_time = timeit.default_timer()

        # Add results to dataset.  Time is saved in milliseconds
        datapoints.append(DataPoint(a, b, gcd, (stop_time - start_time) * 1000))
    return datapoints


"""
Functions to Output Results
"""


def output_to_excel(alg_name, data):
    create_results_xlsx(alg_name, data)
    create_stats_xlsx(alg_name, data)


def create_results_xlsx(alg_name, data):
    # Create the excel file
    workbook = openpyxl.Workbook()
    filename = alg_name + "_Results.xlsx"
    workbook.save(filename)
    workbook = load_workbook(filename)

    # Set the active sheet
    sheet = workbook.active
    sheet.title = alg_name.replace("_", ' ') + ' Results'

    # Format the column headers
    sheet.cell(row=1, column=1).value = 'Number 1'
    sheet.cell(row=1, column=2).value = 'Number 2'
    sheet.cell(row=1, column=3).value = 'GCD'
    sheet.cell(row=1, column=4).value = 'Time (ms)'
    ws = workbook[sheet.title]
    Font(bold=True, italic=True)
    for cell in ws["1:1"]:
        cell.font = Font(bold=True, italic=True)

    # Populate
    row = 1
    for data_point in data:
        row += 1
        sheet.cell(row=row, column=1).value = data_point.value1
        sheet.cell(row=row, column=2).value = data_point.value2
        sheet.cell(row=row, column=3).value = data_point.gcd
        sheet.cell(row=row, column=4).value = data_point.time

    # Set a uniform width
    for column_cells in sheet.columns:
        sheet.column_dimensions[column_cells[0].column].width = 10
    workbook.save(filename)


def create_stats_xlsx(alg_name, data):
    # Create the excel file
    workbook = openpyxl.Workbook()
    filename = alg_name + "_Statistics.xlsx"
    workbook.save(filename)

    # Set the active sheet
    workbook = load_workbook(filename)
    sheet = workbook.active
    sheet.title = alg_name.replace("_", ' ') + ' Stats'

    # Format Header row/columns
    sheet.cell(row=1, column=2).font = Font(bold=True)
    for n in range(1, 6):
        sheet.cell(row=n, column=1).font = Font(bold=True)
    sheet.cell(row=1, column=1).value = 'Statistics'
    sheet.cell(row=1, column=2).value = 'Milliseconds'
    sheet.cell(row=2, column=1).value = 'Maximum'
    sheet.cell(row=3, column=1).value = 'Minimum'
    sheet.cell(row=4, column=1).value = 'Average'
    sheet.cell(row=5, column=1).value = 'Median'

    # Populate
    sheet.cell(row=2, column=2).value = max(data_point.time for data_point in data)  # max time
    sheet.cell(row=3, column=2).value = min(data_point.time for data_point in data)  # min time
    sheet.cell(row=4, column=2).value = sum(data_point.time for data_point in data) / len(data)  # avg time
    sheet.cell(row=5, column=2).value = statistics.median(data_point.time for data_point in data)  # median time

    # Set uniform width
    for column_cells in sheet.columns:
        sheet.column_dimensions[column_cells[0].column].width = 12
    workbook.save(filename)


"""
Helpers
"""


def create_random_array(array_length, rand_min, rand_max):
    rand = []
    for num in range(array_length):
        rand.append(random.randint(rand_min, rand_max))
    return rand


"""
Run Program
"""


if __name__ == "__main__":
    # Array of size 200, with values ranging from 1 to 999
    # Values are used in pairs, so an array of size 200 will produce 100 test runs
    array_size = 200
    rand_min = 1
    rand_max = 999

    init()  # Initialize colorama.  Used for colors in the terminal
    print(colored("\nStarting Euclid Tests\n", 'blue'))

    # Creates a directory on the host in the current path called PA-1-Results, if it did not previously
    # exist.  The output files will be written here.  Any prexisting output files will be overwritten
    if not os.path.exists("PA-1-Results"):
        print("Creating directory PA-1-Results ... ", end="")
        os.mkdir("PA-1-Results")
        print(colored("Done\n", 'green'))
    os.chdir("PA-1-Results")

    print("Creating random inputs from " + str(rand_min) + " to " + str(rand_max) + " ... ", end="")
    numbers = create_random_array(array_size, rand_min, rand_max)
    length = len(numbers) // 2

    print(colored("Done", 'green') + "\nRunning Brute Force ... ", end="")
    output_to_excel("Brute_Force", run_test(brute_force, length, iter(numbers)))

    print(colored("Done", 'green') + "\nRunning Original Euclid ... ", end="")
    output_to_excel("Original_Euclid", run_test(original_euclid, length, iter(numbers)))

    print(colored("Done", 'green') + "\nRunning Second Euclid ... ", end="")
    output_to_excel("Second_Euclid", run_test(second_euclid, length, iter(numbers)))

    print(colored("Done\n", 'green'))
    print(colored("Finished.  ", 'blue') + "Results saved to " + os.getcwd())

    # Added so the console that appears when the executable is run doesn't dissappear automatically
    input("Press Enter to exit ")
